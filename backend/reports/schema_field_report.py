import json
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.common_paths import DATA_DIR

with open(DATA_DIR / 'schema_data.json') as f:
    data = json.load(f)

with open(DATA_DIR / 'binat_data.json') as f:
    binat = json.load(f)

# ── Build direction map from binat ────────────────────────────────────────
# Outbound = SAP is the sender (SAP -> WMS)
# Inbound  = SAP is the receiver (WMS -> SAP)
interface_direction = {}

for b in binat:
    iface_id = str(b['Interface']).strip()
    is_sap_sender   = 'SAP' in str(b.get('sending_system',   '')).upper()
    is_sap_receiver = 'SAP' in str(b.get('receiving_system', '')).upper()
    if is_sap_sender and is_sap_receiver:
        interface_direction[iface_id] = 'Both'
    elif is_sap_sender:
        interface_direction[iface_id] = 'Outbound'
    elif is_sap_receiver:
        interface_direction[iface_id] = 'Inbound'

schemas      = [s['schema'] for s in data]
schema_label = {s['schema']: s['schema'].replace('schema-', '') for s in data}
schema_sheet = {s['schema']: s['sheet_name'] for s in data}

# ── Collect field presence & constraints ──────────────────────────────────
field_data = defaultdict(dict)

for s in data:
    lbl = s['schema']
    for field in s['fields']:
        fname = field.get('json_field', '').strip()
        if not fname:
            continue
        ft  = field.get('field_type', '').strip()
        ln  = field.get('length', '').strip()
        req = field.get('required', '').strip()
        pv  = field.get('possible_values', '').strip()
        parts = []
        if ft: parts.append(ft)
        if ln: parts.append(f'len:{ln}')
        if req.lower() in ('כן', 'yes', 'y', 'true'): parts.append('req')
        if pv and len(pv) <= 30: parts.append(pv)
        field_data[fname][lbl] = ' | '.join(parts)

def consensus(vals):
    non_empty = [v for v in vals if v]
    return Counter(non_empty).most_common(1)[0][0] if non_empty else ''

field_names_sorted = sorted(
    field_data.keys(),
    key=lambda fn: (-len(field_data[fn]), fn)
)

# ── Workbook setup ────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'Field Matrix'

# ── Colours & styles ──────────────────────────────────────────────────────
DARK_BG  = '1F3864'
MED_BG   = '2E75B6'
LIGHT_BG = 'D6E4F0'
WHITE    = 'FFFFFF'
AMBER    = 'FFF2CC'
ABSENT   = 'F2F2F2'
TICK_CLR = '1F7A1F'
OCC_CLR  = '2E4057'

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

hdr_font   = Font(name='Arial', bold=True, color=WHITE, size=9)
field_font = Font(name='Arial', bold=True, size=9)
cons_font  = Font(name='Arial', size=8, color='444444')
tick_font  = Font(name='Arial', bold=True, color=TICK_CLR, size=10)
occ_font   = Font(name='Arial', bold=True, size=9)
body_font  = Font(name='Arial', size=9)

center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left',   vertical='center')

COL_FIELD  = 1
COL_CONS   = 2
COL_OCC    = 3
COL_S0     = 4

# ── Header rows 1 & 2 ────────────────────────────────────────────────────
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 36

for col, label in [(COL_FIELD, 'Field Name'),
                   (COL_CONS,  'Constraints'),
                   (COL_OCC,   'Occurrences')]:
    c = ws.cell(1, col, label)
    c.font = hdr_font
    c.fill = fill(DARK_BG)
    c.alignment = center
    ws.cell(2, col, '').fill = fill(DARK_BG)

for i, schema in enumerate(schemas):
    col = COL_S0 + i
    c1 = ws.cell(1, col, schema_label[schema])
    c1.font = Font(name='Arial', bold=True, color=WHITE, size=8)
    c1.fill = fill(MED_BG)
    c1.alignment = center
    c2 = ws.cell(2, col, schema_sheet[schema])
    c2.font = Font(name='Arial', size=7, color=WHITE)
    c2.fill = fill(DARK_BG)
    c2.alignment = Alignment(horizontal='center', vertical='center', textRotation=90)

# ── Data rows ─────────────────────────────────────────────────────────────
ROW_START = 3
CHECKMARK = '✓'

for row_i, fname in enumerate(field_names_sorted):
    row      = ROW_START + row_i
    presence = field_data[fname]
    count    = len(presence)
    cons     = consensus(list(presence.values()))

    ws.row_dimensions[row].height = 14
    row_bg = AMBER if count == len(schemas) else (LIGHT_BG if row_i % 2 == 0 else WHITE)

    cf = ws.cell(row, COL_FIELD, fname)
    cf.font = field_font; cf.fill = fill(row_bg)
    cf.alignment = left;  cf.border = border

    cc = ws.cell(row, COL_CONS, cons)
    cc.font = cons_font; cc.fill = fill(row_bg)
    cc.alignment = left; cc.border = border

    co = ws.cell(row, COL_OCC, count)
    co.font = Font(name='Arial', bold=True, size=9,
                   color=TICK_CLR if count == len(schemas) else OCC_CLR)
    co.fill = fill(row_bg)
    co.alignment = center; co.border = border

    for i, schema in enumerate(schemas):
        col = COL_S0 + i
        if schema in presence:
            ct = ws.cell(row, col, CHECKMARK)
            ct.font = tick_font
            ct.fill = fill(row_bg)
        else:
            ct = ws.cell(row, col, '')
            ct.fill = fill(ABSENT)
        ct.alignment = center; ct.border = border

# ── Totals row ────────────────────────────────────────────────────────────
total_row = ROW_START + len(field_names_sorted)
ws.row_dimensions[total_row].height = 16

for col in [COL_FIELD, COL_CONS, COL_OCC]:
    ws.cell(total_row, col).fill = fill(DARK_BG)

tl = ws.cell(total_row, COL_FIELD, f'TOTAL  ({len(field_names_sorted)} fields)')
tl.font = Font(name='Arial', bold=True, color=WHITE, size=9)

for i, schema in enumerate(schemas):
    col = COL_S0 + i
    cnt = sum(1 for fn in field_names_sorted if schema in field_data[fn])
    ct  = ws.cell(total_row, col, cnt)
    ct.font = Font(name='Arial', bold=True, color=WHITE, size=8)
    ct.fill = fill(MED_BG)
    ct.alignment = center

# ── Column widths ─────────────────────────────────────────────────────────
ws.column_dimensions[get_column_letter(COL_FIELD)].width = 28
ws.column_dimensions[get_column_letter(COL_CONS)].width  = 30
ws.column_dimensions[get_column_letter(COL_OCC)].width   = 10
for i in range(len(schemas)):
    ws.column_dimensions[get_column_letter(COL_S0 + i)].width = 3.6

ws.freeze_panes = ws.cell(ROW_START, COL_S0)
last_col = get_column_letter(COL_S0 + len(schemas) - 1)
ws.auto_filter.ref = f'A2:{last_col}{total_row - 1}'

# ── Schema Summary sheet ──────────────────────────────────────────────────
ws2 = wb.create_sheet('Schema Summary')
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 28
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 50

# Direction colour coding
DIR_OUT_BG  = 'D9EAD3'   # green tint  – Outbound
DIR_IN_BG   = 'CFE2F3'   # blue tint   – Inbound
DIR_BOTH_BG = 'FFF2CC'   # amber tint  – Both

for ci, h in enumerate(['Schema', 'Sheet(s)', '# Fields', 'Direction', 'Interfaces (SAP)'], 1):
    c = ws2.cell(1, ci, h)
    c.font = hdr_font; c.fill = fill(DARK_BG); c.alignment = center

for ri, s in enumerate(data, 2):
    lbl    = s['schema'].replace('schema-', '')
    ifaces = ', '.join(f"{x['interface']}: {x['description'][:22]}"
                       for x in s['interfaces'][:4])

    # Determine schema-level direction from its SAP interfaces
    dirs = {interface_direction.get(str(x['interface']).strip(), '') for x in s['interfaces']}
    dirs.discard('')
    if len(dirs) == 0:
        direction = ''
    elif len(dirs) > 1 or 'Both' in dirs:
        direction = 'Both'
    else:
        direction = dirs.pop()

    row_bg = LIGHT_BG if ri % 2 == 0 else WHITE
    for ci, val in enumerate([lbl, s['sheet_name'], len(s['fields']), direction, ifaces], 1):
        c = ws2.cell(ri, ci, val)
        if ci == 4:   # Direction column – apply colour coding
            dir_bg = (DIR_OUT_BG  if direction == 'Outbound' else
                      DIR_IN_BG   if direction == 'Inbound'  else
                      DIR_BOTH_BG if direction == 'Both'     else row_bg)
            c.font  = Font(name='Arial', size=9, bold=True)
            c.fill  = fill(dir_bg)
            c.alignment = center
        else:
            c.font  = Font(name='Arial', size=8 if ci == 5 else 9)
            c.fill  = fill(row_bg)
        c.border = border

wb.save('schema_field_report.xlsx')
print(f'Schemas   : {len(schemas)}')
print(f'Interfaces: {sum(len(s["interfaces"]) for s in data)}')
print(f'Fields    : {len(field_names_sorted)} unique  ({len(schemas)} schemas)')